# tests/test_export_xlsx.py
import csv
from openpyxl import load_workbook
from src import export_xlsx


def _build(tmp_path):
    csv_path = tmp_path / "s.csv"
    rows = [["candidate_id", "rank", "score", "reasoning"],
            ["CAND_0000001", "1", "0.9876", "good fit"],
            ["CAND_0000002", "2", "0.9001", "ok fit"]]
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        csv.writer(f).writerows(rows)
    xlsx_path = tmp_path / "s.xlsx"
    export_xlsx.csv_to_xlsx(csv_path, xlsx_path)
    return load_workbook(xlsx_path).active


def test_data_contract_preserved(tmp_path):
    ws = _build(tmp_path)
    got = [[str(c.value) for c in row] for row in ws.iter_rows()]
    # row 1 must be exactly the header; 100->here 2 data rows; ids/order intact
    assert got[0] == ["candidate_id", "rank", "score", "reasoning"]
    assert got[1][0] == "CAND_0000001"
    assert got[2][0] == "CAND_0000002"
    assert len(got) == 3


def test_cells_are_typed(tmp_path):
    ws = _build(tmp_path)
    assert ws["B2"].value == 1 and isinstance(ws["B2"].value, int)       # rank is int
    assert abs(ws["C2"].value - 0.9876) < 1e-9 and isinstance(ws["C2"].value, float)
    assert ws["C2"].number_format == "0.0000"                            # 4-decimal display


def test_visual_styling_applied(tmp_path):
    ws = _build(tmp_path)
    assert ws["A1"].font.bold is True                 # styled header
    assert ws.freeze_panes == "A2"                    # header frozen
    assert ws.auto_filter.ref == "A1:D3"              # filter over the table
    assert ws.column_dimensions["D"].width >= 50      # wide reasoning column
    assert ws["D2"].alignment.wrap_text is True       # reasoning wraps
    # a colour-scale rule is attached to the score column
    rules = [r for rng in ws.conditional_formatting for r in ws.conditional_formatting[rng]]
    assert any(getattr(r, "type", "") == "colorScale" for r in rules)
