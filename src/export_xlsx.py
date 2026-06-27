"""Convert the validator-clean CSV into a polished, presentation-ready XLSX for
the portal upload.

The data contract is preserved exactly: row 1 is the header
``candidate_id,rank,score,reasoning`` and rows 2..N are the ranked candidates in
the same order — so any automated reader parses it identically to the CSV. Only
visual formatting and cell *types* are added (rank as int, score as a 4-decimal
number), which makes the file both nicer to read and cleaner to ingest.
"""
import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

HEADER = ["candidate_id", "rank", "score", "reasoning"]

# palette
_HEADER_FILL = PatternFill("solid", fgColor="0E7C66")     # deep teal
_BAND_FILL = PatternFill("solid", fgColor="F2F7F5")       # very light mint (zebra)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN = Side(style="thin", color="D6DEDB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WIDTHS = {"candidate_id": 16, "rank": 7, "score": 11, "reasoning": 95}


def csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))

    wb = Workbook()
    ws = wb.active
    ws.title = "ranking"

    # --- header row (row 1 — must stay exactly the column header) ---
    for col, name in enumerate(rows[0], start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    # --- data rows (typed + styled) ---
    for i, raw in enumerate(rows[1:], start=2):
        cid, rank_s, score_s, reasoning = (raw + ["", "", "", ""])[:4]
        band = _BAND_FILL if i % 2 == 0 else None

        c_id = ws.cell(row=i, column=1, value=cid)
        c_id.font = Font(name="Consolas", size=10)
        c_id.alignment = Alignment(horizontal="center", vertical="center")

        c_rank = ws.cell(row=i, column=2, value=int(rank_s) if rank_s.strip() else None)
        c_rank.alignment = Alignment(horizontal="center", vertical="center")

        c_score = ws.cell(row=i, column=3,
                          value=float(score_s) if score_s.strip() else None)
        c_score.number_format = "0.0000"
        c_score.alignment = Alignment(horizontal="center", vertical="center")

        c_reason = ws.cell(row=i, column=4, value=reasoning)
        c_reason.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        for col in range(1, 5):
            cell = ws.cell(row=i, column=col)
            cell.border = _BORDER
            if band and col != 3:        # leave the score column for the colour scale
                cell.fill = band
        ws.row_dimensions[i].height = 30

    n = len(rows) - 1  # number of data rows

    # column widths + header height
    for col, name in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(col)].width = _WIDTHS.get(name, 14)
    ws.row_dimensions[1].height = 22

    # freeze the header, add an auto-filter, and a red->yellow->green gradient on score
    ws.freeze_panes = "A2"
    if n > 0:
        ws.auto_filter.ref = f"A1:D{n + 1}"
        ws.conditional_formatting.add(
            f"C2:C{n + 1}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",                 # low = red
                mid_type="percentile", mid_value=50, mid_color="FFEB84", # mid = yellow
                end_type="max", end_color="63BE7B",                      # high = green
            ),
        )

    wb.save(xlsx_path)


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args(argv)
    csv_to_xlsx(Path(args.csv), Path(args.out))


if __name__ == "__main__":
    main()
